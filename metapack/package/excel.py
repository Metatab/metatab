# Copyright (c) 2017 Civic Knowledge. This file is licensed under the terms of the
# MIT License, included in this distribution as LICENSE

""" """

from os.path import join, dirname
from .core import PackageBuilder
from metapack.exc import PackageError
from metapack.util import  ensure_dir, slugify
from itertools import islice


class ExcelPackageBuilder(PackageBuilder):
    """An Excel File Package"""

    def __init__(self, source_ref=None, package_root=None,  callback=None, env=None):
        super().__init__(source_ref, package_root, callback, env)

        self.cache_path = self.package_name+".xlsx"

        self.package_path = self.package_root.join(self.cache_path)

        self.package_root.ensure_dir()


    def save(self):
        from openpyxl import Workbook
        from openpyxl.cell import WriteOnlyCell
        from openpyxl.styles import PatternFill, Alignment

        self.check_is_ready()

        self.wb = Workbook(write_only=True)

        meta_ws = self.wb.create_sheet()
        meta_ws.title = "meta"
        meta_ws.sheet_properties.tabColor = "8888ff"

        meta_ws.column_dimensions['A'].width = 15
        meta_ws.column_dimensions['B'].width = 40
        meta_ws.column_dimensions['C'].width = 20
        meta_ws.column_dimensions['D'].width = 20
        meta_ws.column_dimensions['E'].width = 20

        self.sections.resources.sort_by_term()

        self.load_declares()

        self.doc.cleanse()

        self._load_resources()

        self._clean_doc()

        fill = PatternFill("solid", fgColor="acc0e0")  # PatternFill(patternType='gray125')
        table_fill = PatternFill("solid", fgColor="d9dce0")  # PatternFill(patternType='gray125')

        alignment = Alignment(wrap_text=False)
        for i, row in enumerate(self.doc.rows, 1):

            if row[0] == 'Section' or row[0] == 'Table':
                styled_row = []
                for c in row + [''] * 5:
                    cell = WriteOnlyCell(meta_ws, value=c)
                    cell.fill = fill if row[0] == 'Section' else table_fill
                    styled_row.append(cell)
                meta_ws.append(styled_row)


            else:
                meta_ws.append(row)


        self.wb.save(self.package_path.path)

        return self.package_path

    def _load_resources(self):
        """Remove the geography from the files, since it isn't particularly useful in Excel"""

        for t in self.doc.find('Root.Table'):
            for c in t.find('Table.Column'):
                if c.get_value('datatype') == 'geometry':
                    c['transform'] = '^empty_str'
                    c['datatype'] = 'text'

        return super()._load_resources()

    def _load_resource(self, source_r):

        r = self._doc.resource(source_r.name)

        from itertools import islice

        self.prt("Loading data for sheet '{}' ".format(r.name))

        ws = self.wb.create_sheet(r.name)

        ws.append(r.headers)


        for row in islice(r, 1, None):
            ws.append(row)

        r.url = r.name