# Copyright (c) 2016 Civic Knowledge. This file is licensed under the terms of the
# Revised BSD License, included in this distribution as LICENSE

"""
A Class for writing Excel packages in Metatab package format.
"""



from collections import namedtuple

TableColumn = namedtuple('TableColumn','path name start_line header_lines columns')

class ExcelPackage(object):

    def __init__(self, path):

        self.path = path

        self.init_wb()
        self.init_doc()

    def init_wb(self):
        from openpyxl import Workbook

        wb = Workbook(write_only=True)
        ws = wb.create_sheet()
        ws.title = "meta"
        ws.sheet_properties.tabColor = "8888ff"

        self.wb = wb
        self.meta_ws = ws

        return wb

    def copy_section(self, section_name, doc):

        for t in doc[section_name]:
            self.doc.add_term(t)

    @staticmethod
    def extract_path_name(ref):
        from os.path import splitext, basename, abspath
        from rowgenerators.util import parse_url_to_dict

        uparts = parse_url_to_dict(ref)

        if not uparts['scheme']:
            path = abspath(ref)
            name = basename(splitext(path)[0])
            ref = "file://"+path
        else:
            path = ref
            name = basename(splitext(uparts['path'])[0])

        return ref, path, name

    def add_data_file(self, ref, name, description, columns, start_line, encoding='latin1', cache=None):
        from rowgenerators import RowGenerator
        from fs.opener import fsopendir
        from itertools import islice
        self.resources.new_term('Datafile', name,  description=description)

        if cache is None:
            cache = fsopendir('/tmp')

        table = self.schema.new_term('Table', name)

        ref, path, name = self.extract_path_name(ref)

        for c in columns:
            table.new_child('Column', c['name'], datatype=c['datatype'])

        gen = islice(RowGenerator(url=ref, cache=cache, encoding=encoding),start_line, None)

        ws = self.wb.create_sheet()
        ws.title = name
        #ws.sheet_properties.tabColor = "8888ff"

        ws.append([c['name'] for c in columns])

        for row in gen:
            ws.append(row)


    def init_doc(self):

        from metatab import MetatabDoc
        from os.path import join

        self.doc = MetatabDoc()

        self.resources = self.doc.get_or_new_section('Resources', ['Description'])

        self.schema = self.doc.get_or_new_section('Schema',
                                                  ['DataType', 'AltName', 'Description'])

        return self.doc

    def save(self):

        for row in self.doc.rows:
            self.meta_ws.append(row)

        self.wb.save(self.path)