# Copyright (c) 2016 Civic Knowledge. This file is licensed under the terms of the
# MIT License, included in this distribution as LICENSE

"""Class for writing Metapack packages"""

from __future__ import print_function
import json
import shutil
from collections import namedtuple
from io import BytesIO
from itertools import islice
from os import getcwd, makedirs, remove, walk
from os.path import basename, abspath, dirname, exists, isdir, join, splitext

import unicodecsv as csv
from six import string_types, text_type

from metatab import TermParser, MetatabDoc, DEFAULT_METATAB_FILE, Resource
from metatab.datapackage import convert_to_datapackage
from metatab.util import slugify, Bunch
from rowgenerators import RowGenerator, SourceSpec, TextEncodingError, Url, enumerate_contents
from rowgenerators.util import get_cache
from tableintuit import RowIntuiter
from .exc import PackageError
from rowgenerators.generators import get_dflo, download_and_cache
from rowgenerators import SourceSpec
from rowgenerators.exceptions import DownloadError
from os.path import basename, splitext

TableColumn = namedtuple('TableColumn', 'path name start_line header_lines columns')


def ensure_exists(d):
    from os import makedirs
    from os.path import exists

    if not exists(d):
        makedirs(d)


def write_csv(path_or_flo, headers, gen):
    try:
        f = open(path_or_flo, "wb")

    except TypeError:
        f = path_or_flo  # Assume that it's already a file-like-object

    try:
        w = csv.writer(f)
        w.writerow(headers)

        row = None
        try:
            for row in gen:
                w.writerow(row)
        except:
            import sys
            print("write_csv: ERROR IN ROW", row, file=sys.stderr)
            raise

        try:
            return f.getvalue()
        except AttributeError:
            return None

    finally:
        f.close()


def write_geojson(path_or_flo, columns, gen):
    import fiona
    from fiona.crs import from_epsg
    from collections import OrderedDict

    type_map = {
        'number': 'float'
    }

    schema = {
        'geometry': 'Unknown',
        'properties': OrderedDict(
            [(c['header'], type_map.get(c['datatype'], c['datatype'])) for c in columns]
        )
    }

    try:
        f = fiona.open(path_or_flo,
                       driver='GeoJSON',
                       crs=from_epsg(4326),
                       schema=schema)
    except TypeError:
        f = path_or_flo  # Assume that it's already a file-like-object

    try:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(gen)

        try:
            return f.getvalue()
        except AttributeError:
            return None

    finally:
        f.close()


class Package(object):
    def __new__(cls, ref=None, cache=None, callback=None, env=None, save_url=None, acl=None):

        if cls == Package:

            if isinstance(ref, Url):
                b = Bunch(ref.dict)
            else:
                b = Bunch(Url(ref).dict)

            if b.resource_format in ('xls', 'xlsx'):
                return super(Package, cls).__new__(ExcelPackage)
            elif b.resource_format == 'zip':
                return super(Package, cls).__new__(ZipPackage)
            elif b.proto == 'gs':
                return super(Package, cls).__new__(GooglePackage)
            elif b.proto == 's3':
                return super(Package, cls).__new__(S3Package)
            elif b.resource_format == 'csv' or b.target_format == 'csv':
                return super(Package, cls).__new__(CsvPackage)
            else:
                raise PackageError("Can't determine package type for ref '{}' ".format(ref))

        else:
            return super(Package, cls).__new__(cls)

    def __init__(self, ref=None, cache=None, callback=None, env=None):

        self._cache = cache if cache else get_cache('metapack')
        self._ref = ref
        self._doc = None
        self._callback = callback
        self._env = env if env is not None else {}
        self.source_dir = dirname(Url(ref).path())
        self.init_doc()

    def load_doc(self, ref):

        if isinstance(ref, string_types):
            self._doc = MetatabDoc(ref, cache=self._cache)
        else:
            self._doc = ref

        return self

    def init_doc(self):

        if self._ref:
            self.load_doc(self._ref)
        else:
            self._doc = MetatabDoc()

            if not self._doc.find("Root.Declare"):
                # FIXME. SHould really have a way to insert this term as the first term.
                self.sections.root.new_term('Declare', 'metatab-latest')
                self._doc.load_declarations(['metatab-latest'])

        return self.doc

    def _open(self):
        """Open the package, possibly downloading it to the cache."""
        if not self._cache:
            raise IOError(
                "Package must have a cache, set either in the package constructor, or with the METAPACK_CACHE env var")

    @property
    def path(self):
        return self._ref

    def save_path(self):
        """Default path for the file to be wrotten to"""
        raise NotImplementedError()

    def exists(self, path=None):
        return exists(self.save_path(path))

    @property
    def doc(self):
        """Return the Metatab metadata document"""
        if not self._doc and self._ref:
            self._doc = MetatabDoc(TermParser(self._ref))

        return self._doc

    def copy_section(self, section_name, doc):

        for t in doc[section_name]:
            self.doc.add_term(t)

    def resources(self, name=None, term=None, section='resources'):
        for r in self.doc.resources(name, term, section=section):
            yield r

    @property
    def datafiles(self):

        for r in self.doc.resources(term=['root.datafile', 'root.suplimentarydata', 'root.datadictionary']):
            yield r

    def datafile(self, ref):
        """Return a resource, as a file-like-object, given it's name or url as a reference. """

        r = list(self.doc.resources(name=ref))

        if not r:
            return None
        else:
            return r[0]

    @property
    def documentation(self):
        for t in self.doc.terms:
            if t.term_is(['root.documentation']):
                yield Resource(t, self)

    def schema(self, ref):
        """Return information about a resource, given a name or url"""
        raise NotImplementedError()

    @property
    def package_name(self):
        return self._doc.find_first_value('Root.Name')

    @property
    def sections(self):

        class _Sections(object):

            def __init__(self, doc):
                self.doc = doc

            @property
            def root(self):
                return self.doc['root']

            @property
            def resources(self):
                if not 'resources' in self.doc:
                    self.doc.get_or_new_section('Resources',
                                                "Name Schema Space Time StartLine HeaderLines Encoding Description".split())

                return self.doc['resources']

            @property
            def contacts(self):
                if not 'Contacts' in self.doc:
                    self.doc.get_or_new_section('Contacts', 'Email Org Url'.split())

                return self.doc['Contacts']

            @property
            def documentation(self):
                if not 'Documentation' in self.doc:
                    self.doc.get_or_new_section('Documentation', 'Title Schema Description '.split())

                return self.doc['documentation']

            @property
            def schema(self):
                if not 'Schema' in self.doc:
                    self.doc.get_or_new_section('Schema', 'DataType AltName Description'.split())

                return self.doc['schema']

        return _Sections(self._doc)

    @staticmethod
    def extract_path_name(ref):

        du = Url(ref)

        if du.proto == 'file':
            path = abspath(ref)
            name = basename(splitext(path)[0])
            ref = "file://" + path
        else:
            path = ref

            if du.target_segment:
                try:
                    int(du.target_segment)
                    name = du.target_file + text_type(du.target_segment)

                except ValueError:
                    name = du.target_segment

            else:
                name = splitext(du.target_file)[0]

        return ref, path, name

    @staticmethod
    def classify_url(url):

        ss = SourceSpec(url=url)

        if ss.format in ('xls', 'xlsx', 'tsv', 'csv'):
            term_name = 'DataFile'
        elif ss.format in ('pdf', 'doc', 'docx', 'html'):
            term_name = 'Documentation'
        else:
            term_name = 'Resource'

        return term_name

    @staticmethod
    def run_row_intuit(path, cache):

        for encoding in ('ascii', 'utf8', 'latin1'):
            try:
                rows = list(islice(RowGenerator(url=path, encoding=encoding, cache=cache), 5000))
                return encoding, RowIntuiter().run(list(rows))
            except TextEncodingError:
                pass

        raise Exception('Failed to convert with any encoding')

    @staticmethod
    def find_files(base_path, types):

        for root, dirs, files in walk(base_path):
            if '_metapack' in root:
                continue

            for f in files:
                if f.startswith('_'):
                    continue

                b, ext = splitext(f)
                if ext[1:] in types:
                    yield join(root, f)

    def load_declares(self):

        self.doc.load_declarations(t.value for t in self.doc.find('Root.Declare'))

    def prt(self, *args):
        if self._callback:
            self._callback(*args)

    def warn(self, *args):
        if self._callback:
            self._callback('WARN', *args)

    def err(self, *args):
        if self._callback:
            self._callback('ERROR', *args)

    def add_single_resource(self, ref, **properties):
        """ Add a single resource, without trying to enumerate it's contents
        :param ref:
        :return:
        """

        t = self.doc.find_first('Root.Datafile', value=ref)

        if t:
            self.prt("Datafile exists for '{}', deleting".format(ref))
            self.doc.remove_term(t)

        term_name = self.classify_url(ref)

        ref, path, name = self.extract_path_name(ref)

        self.prt("Adding resource for '{}'".format(ref))

        try:
            encoding, ri = self.run_row_intuit(path, self._cache)
        except Exception as e:
            self.warn("Failed to intuit '{}'; {}".format(path, e))
            return None

        if not name:
            from hashlib import sha1
            name = sha1(slugify(path).encode('ascii')).hexdigest()[:12]

            # xlrd gets grouchy if the name doesn't start with a char
            try:
                int(name[0])
                name = 'a' + name[1:]
            except:
                pass

        if 'name' in properties:
            name = properties['name']
            del properties['name']

        return self.sections.resources.new_term(term_name, ref, name=name,
                                                startline=ri.start_line,
                                                headerlines=','.join(str(e) for e in ri.header_lines),
                                                encoding=encoding,
                                                **properties)

    def add_resource(self, ref, **properties):
        """Add one or more resources entities, from a url and property values,
        possibly adding multiple entries for an excel spreadsheet or ZIP file"""

        raise NotImplementedError("Still uses decompose_url")

        du = Bunch(decompose_url(ref))

        added = []
        if du.proto == 'file' and isdir(ref):
            for f in self.find_files(ref, ['csv']):

                if f.endswith(DEFAULT_METATAB_FILE):
                    continue

                if self._doc.find_first('Root.Datafile', value=f):
                    self.prt("Datafile exists for '{}', ignoring".format(f))
                else:
                    added.extend(self.add_resource(f, **properties))
        else:
            self.prt("Enumerating '{}'".format(ref))

            for c in enumerate_contents(ref, self._cache):
                added.append(self.add_single_resource(c.rebuild_url(), **properties))

        return added

    def _clean_doc(self, doc=None):
        """Clean the doc before writing it, removing unnecessary properties and doing other operations."""

        if doc is None:
            doc = self.doc

        resources = doc['Resources']

        # We don't need these anymore because all of the data written into the package is normalized.
        for arg in ['startline', 'headerlines', 'encoding']:
            for e in list(resources.args):
                if e.lower() == arg:
                    resources.args.remove(e)

        for term in resources:
            term['startline'] = None
            term['headerlines'] = None
            term['encoding'] = None

        schema = doc['Schema']

        ## FIXME! This is probably dangerous, because the section args are changing, but the children
        ## are not, so when these two are combined in the Term.properties() acessors, the values are off.
        ## Because of this, _clean_doc should be run immediately before writing the doc.
        for arg in ['altname', 'transform']:
            for e in list(schema.args):
                if e.lower() == arg:
                    schema.args.remove(e)

        for table in self.doc.find('Root.Table'):
            for col in table.find('Column'):
                try:
                    col.value = col['altname'].value
                except:
                    pass

                col['altname'] = None
                col['transform'] = None

        return doc

    def _load_resources(self):
        """Copy all of the Datafile entries into the package"""

        for r in self.datafiles:

            if not r.url:
                self.warn("No value for URL for {} ".format(r.term))
                continue

            try:
                if self._resource.exists(r):
                    self.prt("Resource '{}' exists, skipping".format(r.name))
                continue
            except AttributeError:
                pass

            self.prt("Reading resource {} from {} ".format(r.name, r.resolved_url))

            if not r.headers:
                raise PackageError("Resource {} does not have header. Have schemas been generated?".format(r.name))

            try:
                code_path = join(dirname(self.package_dir), 'row_processors', r.name + '.py')
            except AttributeError:
                # Not all packages have a package_dir property, which is an error,
                # but we really only need code_path for the FilesystemPackage
                code_path = None


            rg = self.doc.resource(r.name, env=self._env,code_path=code_path)

            assert rg is not None

            if True:
                # Skipping the first line because we'll insetrt the headers manually
                self._load_resource(r, islice(rg, 1, None), r.headers)

            else:

                # Old code; no longer sure why the code is faking the start line and removing
                # the format, headerlines, etc.

                start_line = int(r.get_value('startline')) if r.get_value('startline') is not None  else 1

                gen = islice(rg, start_line, None)

                r.encoding = None
                r.startline = None
                r.headerlines = None
                r.format = None

                self._load_resource(r, gen, r.headers)

    def _get_ref_contents(self, t):

        uv = Url(t.value)
        ur = Url(self._ref)

        # In the case that the input doc is a file, and the ref is to a file,
        # try interpreting the file as relative.
        if ur.proto == 'file' and uv.proto == 'file':
            path = uv.prefix_path(Url(ur.dirname()).parts.path)
            ss = SourceSpec(path)

        else:
            ss = SourceSpec(t.value)

        return self._download_ss(ss)

    def _download_ss(self,ss):

        try:
            d = download_and_cache(ss, self._cache)
        except DownloadError as e:
            self.warn("Failed to load file for '{}': {}".format(ss, e))
            return None, None

        dflo = get_dflo(ss, d['sys_path'])

        f = dflo.open('rb')

        try:
            # For file file, the target_file may actually be a regex, so we have to resolve the
            # regex before using it as a filename
            real_name = basename(dflo.memo[1].name)  # Internal detail of how Zip files are accessed
        except (AttributeError, TypeError):

            real_name = basename(ss.target_file)

        return real_name, f

    def _load_documentation_files(self):
        """Copy all of the Datafile entries into the Excel file"""

        for doc in self.doc.find(['Root.Documentation', 'Root.Image']):

            real_name, f = self._get_ref_contents(doc)

            if not f:
                continue

            if doc.term_is('Root.Documentation'):
                # Prefer the slugified title to the base name, because in cases of collections
                # of many data releases, like annual datasets, documentation files may all have the same name,
                # but the titles should be different.
                real_name_base, ext = splitext(real_name)

                name = doc.get_value('name') if doc.get_value('name') else real_name_base
                real_name = slugify(name) + ext

            self._load_documentation(doc, f.read(), real_name)

            f.close()

    def _load_documentation(self, term, contents):
        raise NotImplementedError()

    def _load_files(self):
        """Load other files"""

        def copy_dir(path):
            for (dr, _, files) in walk(path):
                for fn in files:

                    if '__pycache__' in fn:
                        continue

                    relpath = dr.replace(self.source_dir, '').strip('/')
                    src = join(dr, fn)
                    dest = join(relpath, fn)

                    real_name, f = self._download_ss(SourceSpec(src))

                    self._load_file( dest, f.read())


        for term in self.resources(term = 'Root.Pythonlib'):

            uv = Url(term.value)
            ur = Url(self._ref)

            # In the case that the input doc is a file, and the ref is to a file,
            # try interpreting the file as relative.
            if ur.proto == 'file' and uv.proto == 'file':

                # Either a file or a directory
                path = join(self.source_dir, uv.path())
                if isdir(path):
                    copy_dir(path)

            else:
                # Load it as a URL
                real_name, f = self._get_ref_contents(term)
                self._load_file(term.value,f.read() )

        nb_dir = join(self.source_dir, 'notebooks')

        if exists(nb_dir) and isdir(nb_dir):
            copy_dir(nb_dir)



    def _load_file(self,  filename, contents):
        raise NotImplementedError()


    def check_is_ready(self):
        pass


class FileSystemPackage(Package):
    """A File System package"""

    def __init__(self, path=None, callback=None, cache=None, env=None):

        super(FileSystemPackage, self).__init__(path, callback=callback, cache=cache, env=env)
        self.package_dir = None

    def exists(self, path=None):

        if self.package_dir is None:
            self._init_dir(path)

        return exists(join(self.save_path(path), DEFAULT_METATAB_FILE))

    def is_older_than_metatada(self, path=None):
        """
        Return True if the package save file is older than the metadata. Returns False if the time of either can't be determined

        :param path: Optional extra save path, used in save_path()

        """

        from os.path import getmtime

        try:
            return getmtime(self.save_path(path) + "/metadata.csv") > self._doc.mtime
        except (FileNotFoundError, OSError):
            return False

    def save_path(self, path=None):

        base = self.doc.find_first_value('Root.Name')

        if path and not path.endswith('.zip'):
            return join(path, base)
        else:
            return base

    def save(self, path=None):

        self.check_is_ready()

        if not self.doc.find_first_value('Root.Name'):
            raise PackageError("Package must have Root.Name term defined")

        self.sections.resources.sort_by_term()

        self.doc.cleanse()

        self.load_declares()

        self._init_dir(path)

        ensure_exists(dirname(self.save_path(path)))

        self._load_documentation_files()

        self._load_resources()

        self._load_files()

        self._write_dpj()

        self._clean_doc()

        doc_file = self._write_doc()

        self._write_html()

        return doc_file

    def remove(self, path=None):

        if path is None:
            path = getcwd()

        name = self.doc.find_first_value('Root.Name')

        np = join(path, name)

        if isdir(np):
            shutil.rmtree(np)

    def _init_dir(self, path=None):

        if path is None:
            path = getcwd()

        name = self.doc.find_first_value('Root.Name')

        assert path
        assert name

        np = join(path, name)

        if not isdir(np):
            makedirs(np)

        self.package_dir = np

    def _write_doc(self):
        path = join(self.package_dir, DEFAULT_METATAB_FILE)
        self._doc.write_csv(path)
        return path

    def _write_dpj(self):

        with open(join(self.package_dir, 'datapackage.json'), 'w') as f:
            f.write(json.dumps(convert_to_datapackage(self._doc), indent=4))

    def _write_html(self):

        with open(join(self.package_dir, 'index.html'), 'w') as f:
            f.write(self._doc.html)

    def _load_resource(self, r, gen, headers):

        self.prt("Loading data for '{}' ".format(r.name))

        r.url = 'data/' + r.name + '.csv'

        path = join(self.package_dir, r.url)

        makedirs(dirname(path), exist_ok=True)

        if exists(path):
            remove(path)

        write_csv(path, headers, gen)

        # Writting between resources so row-generating programs and notebooks can
        # access previously created resources. We have to clean the doc before writing it

        ref = self._write_doc()

        # What a wreck ... we also have to get rid of the 'Transform' values, since the CSV files
        # that are written don't need them, and a lot of intermediate processsing ( specifically,
        # jupyter Notebooks, ) does not load them.
        p = FileSystemPackage(ref)
        p._init_dir('_packages')
        p._clean_doc()
        ref = p._write_doc()

    def _load_documentation_files(self):
        from nbconvert.writers import FilesWriter
        from metatab.jupyter.exporters import DocumentationExporter

        notebook_docs = []

        # First find and remove them from the doc.
        for term in list(self.doc['Documentation'].find('Root.Documentation')):
            u = Url(term.value)
            if u.target_format == 'ipynb':
                notebook_docs.append(term)
                self.doc.remove_term(term)

        # Process all of the normal files
        super()._load_documentation_files()

        de = DocumentationExporter()
        fw = FilesWriter()
        fw.build_directory = join(self.package_dir,'docs')

        # Now, generate the documents directly into the filesystem package
        for term in notebook_docs:
            u = Url(term.value)
            nb_path = u.path(self.source_dir)

            output, resources = de.from_filename(nb_path)
            fw.write(output, resources, notebook_name='notebook')

            de.update_metatab(self.doc, resources)

    def _load_documentation(self, term, contents, file_name):

        try:
            title = term['title'].value
        except KeyError:
            self.warn("Documentation has no title, skipping: '{}' ".format(term.value))
            return

        self.prt("Loading documentation for '{}', '{}' ".format(title, file_name))

        path = join(self.package_dir, 'docs/' + file_name)

        makedirs(dirname(path), exist_ok=True)

        if exists(path):
            remove(path)

        with open(path, 'wb') as f:
            f.write(contents)

    def _load_file(self,  filename, contents):

        from metatab.util import ensure_dir

        if "__pycache__" in filename:
            return

        path = join(self.package_dir, filename)

        ensure_dir(dirname(path))

        with open(path, 'wb') as f:
            f.write(contents)


class CsvPackage(Package):
    """"""

    def save_path(self, path=None):
        base = self.doc.find_first_value('Root.Name') + '.csv'

        if path is None:
            path = getcwd()

        if path and not path.endswith('.csv'):
            return join(path, base)
        elif path:
            return path
        else:
            return base

    def _load_resource(self, r, gen, headers):

        r.url = r.resolved_url


    def _relink_documentation(self):

        for doc in self.doc.resources(term=['Root.Documentation', 'Root.Image'], section='Documentation'):
            doc.url =  doc.resolved_url

    def save(self, path=None):

        # HACK ...
        if not self.doc.ref:
            self.doc._ref = self.save_path(path)  # Really should not do this but ...

        self.check_is_ready()

        self.load_declares()

        self.doc.cleanse()

        self._load_resources()

        self._relink_documentation()

        self._clean_doc()

        _path = self.save_path(path)

        ensure_exists(dirname(_path))

        self.doc.write_csv(_path)

        return _path


class ExcelPackage(Package):
    """An Excel File Package"""

    def __init__(self, ref=None, callback=None, cache=None, env=None):

        super(ExcelPackage, self).__init__(ref, callback=callback, cache=cache, env=env)

    def save_path(self, path=None):
        base = self.doc.find_first_value('Root.Name') + '.xlsx'

        if path and not path.endswith('.xlsx'):
            return join(path, base)
        elif path:
            return path
        else:
            return base

    def save(self, path=None):
        from openpyxl import Workbook
        from openpyxl.cell import WriteOnlyCell
        from openpyxl.styles import PatternFill, Font, Alignment

        self.check_is_ready()

        self.wb = Workbook(write_only=True)

        meta_ws = self.wb.create_sheet()
        meta_ws.title = "meta"
        meta_ws.sheet_properties.tabColor = "8888ff"

        meta_ws.column_dimensions['A'].width = 15
        meta_ws.column_dimensions['B'].width = 40
        meta_ws.column_dimensions['C'].width = 20
        meta_ws.column_dimensions['D'].width = 20
        meta_ws.column_dimensions['E'].width = 20

        if not self.doc.find_first_value('Root.Name'):
            raise PackageError("Package must have Root.Name term defined")

        self.sections.resources.sort_by_term()

        self.load_declares()

        self.doc.cleanse()

        self._load_resources()

        self._clean_doc()

        fill = PatternFill("solid", fgColor="acc0e0")  # PatternFill(patternType='gray125')
        table_fill = PatternFill("solid", fgColor="d9dce0")  # PatternFill(patternType='gray125')

        alignment = Alignment(wrap_text=False)
        for i, row in enumerate(self.doc.rows, 1):

            if row[0] == 'Section' or row[0] == 'Table':
                styled_row = []
                for c in row + [''] * 5:
                    cell = WriteOnlyCell(meta_ws, value=c)
                    cell.fill = fill if row[0] == 'Section' else table_fill
                    styled_row.append(cell)
                meta_ws.append(styled_row)


            else:
                meta_ws.append(row)

        ensure_exists(dirname(self.save_path(path)))

        self.wb.save(self.save_path(path))

        return self.save_path(path)

    def _load_resources(self):
        """Remove the geography from the files, since it isn't particularly useful in Excel"""

        for t in self.doc.find('Root.Table'):
            for c in t.find('Table.Column'):
                if c.get_value('datatype') == 'geometry':
                    c['transform'] = '^empty_str'
                    c['datatype'] = 'text'

        return super()._load_resources()

    def _load_resource(self, r, gen, headers):

        self.prt("Loading data for sheet '{}' ".format(r.name))

        ws = self.wb.create_sheet(r.name)

        r.url = r.name

        ws.append(headers)

        for row in gen:
            ws.append(row)


class ZipPackage(Package):
    """A Zip File package"""

    def __init__(self, ref=None, callback=None, cache=None, env=None):

        super(ZipPackage, self).__init__(ref, callback=callback, cache=cache, env=env)

    def save_path(self, path=None):
        base = self.doc.find_first_value('Root.Name') + '.zip'

        if path and not path.endswith('.zip'):
            return join(path, base)
        elif path:
            return path
        else:
            return base

    def _init_zf(self, path):

        from zipfile import ZipFile

        self.zf = ZipFile(self.save_path(path), 'w')

    def save(self, path=None):

        self.check_is_ready()

        root_dir = self.doc.find_first_value('Root.Name')

        self._init_zf(path)

        for root, dirs, files in walk(self.source_dir):
            for f in files:
                source = join(root, f)
                rel = source.replace(self.source_dir,'').strip('/')
                dest = join(root_dir, rel)

                self.zf.write(source,dest)

        self.zf.close()

        return self.save_path(path)


class S3Package(Package):
    """A Zip File package"""

    def __init__(self, path=None, callback=None, cache=None, env=None, save_url=None, acl=None, force=False):

        super(S3Package, self).__init__(path, callback=callback, cache=cache, env=env)

        self._save_url = save_url

        self.bucket = None

        self.force = force

        self._acl = acl if acl else 'public-read'

    def _init_bucket(self, url=None, acl=None):

        from metatab.s3 import S3Bucket

        self._acl = acl if acl is not None else self._acl

        if not self.bucket:
            url = url or self._save_url

            self.bucket = S3Bucket(url, acl=acl)

    def save(self, url=None, acl=None):

        self.check_is_ready()

        self._init_bucket(url, acl)

        name = self.doc.find_first_value('Root.Name')

        if not name:
            raise PackageError("Package must have Root.Name term defined")

        self.prt("Preparing S3 package '{}' ".format(name))

        self.sections.resources.sort_by_term()

        self.doc.cleanse()

        self.load_declares()

        self._load_documentation_files()

        self._load_resources()

        self._clean_doc()

        self._write_doc()

        self._write_dpj()

        self._write_html()

        self.close()

        return self.access_url

    def close(self):
        pass

    @property
    def access_url(self):
        import boto3

        self._init_bucket()  # Fervently hope that the self.save_url has been set

        return self.bucket.access_url(self.package_name, DEFAULT_METATAB_FILE)

    @property
    def private_access_url(self):
        import boto3

        self._init_bucket()  # Fervently hope that the self.save_url has been set

        return self.bucket.private_access_url(self.package_name, DEFAULT_METATAB_FILE)

    @property
    def public_access_url(self):
        import boto3

        self._init_bucket()  # Fervently hope that the self.save_url has been set

        return self.bucket.public_access_url(self.package_name, DEFAULT_METATAB_FILE)

    @property
    def signed_url(self):
        """A URL with an access signature or password """
        import boto3

        self._init_bucket()  # Fervently hope that the self.save_url has been set

        return self.bucket.signed_access_url(self.package_name, DEFAULT_METATAB_FILE)

    def exists(self, url=None):
        import botocore
        from .s3 import S3Bucket

        self._init_bucket(url)

        return self.bucket.exists(self.package_name, 'index.html')

    def write_to_s3(self, path, body):

        self.bucket.write(body, join(self.package_name, path), acl=self._acl, force=self.force)

        return

    def _write_doc(self):

        bio = BytesIO()
        writer = csv.writer(bio)
        writer.writerows(self.doc.rows)

        self.write_to_s3('metadata.csv', bio.getvalue())

    def _write_dpj(self):

        self.write_to_s3('datapackage.json', json.dumps(convert_to_datapackage(self._doc), indent=4))

    def _write_html(self):
        old_ref = self._doc._ref
        self._doc._ref = self.access_url + '/metatab.csv'
        self.write_to_s3('index.html', self._doc.html)
        self._doc._ref = old_ref

    def _load_resource(self, r, gen, headers):

        r.url = 'data/' + r.name + '.csv'

        bio = BytesIO()

        data = write_csv(bio, headers, gen)

        self.prt("Loading data ({} bytes) to '{}' ".format(len(data), r.url))

        self.write_to_s3(r.url, data)

    def _load_documentation(self, term, contents, file_name):

        title = term['title'].value

        self.prt("Loading documentation for '{}', '{}' ".format(title, file_name))

        term['url'].value = 'docs/' + file_name

        self.write_to_s3(term['url'].value, contents)
